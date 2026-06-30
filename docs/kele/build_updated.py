"""
Generates Kele_Contents_BAPI_Updated.xlsx with all bapihvac.com URLs
replaced by SharePoint URLs.

SharePoint base: https://bapisensors.sharepoint.com/sites/BAPIMediaLibrary/Shared Documents/
Folders: Datasheets / Instructions / Product Images
"""

import zipfile, re, xml.etree.ElementTree as ET
import openpyxl
from urllib.parse import quote

# ---- Config ----
SP_BASE = "https://bapisensors.sharepoint.com/sites/BAPIMediaLibrary/Shared%20Documents"
SP_FOLDERS = {
    'datasheets': f"{SP_BASE}/Datasheets",
    'instructions': f"{SP_BASE}/Instructions",
    'images': f"{SP_BASE}/Product%20Images",
}

# ---- Helpers ----
def get_shared_strings(z):
    with z.open('xl/sharedStrings.xml') as f:
        content = f.read().decode('utf-8')
    return re.findall(r'<t[^>]*>(.*?)</t>', content, re.DOTALL)

def col_letter_to_index(col_str):
    """Convert Excel column letter(s) like 'A', 'B', 'AA' to 0-based index."""
    result = 0
    for ch in col_str:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1

def get_sheet_data(z, strings):
    with z.open('xl/worksheets/sheet1.xml') as f:
        tree = ET.parse(f)
    ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    root = tree.getroot()
    rows = []
    for row in root.findall('.//s:row', ns):
        # Use cell reference (e.g. 'A1', 'C3') to place values in the correct column
        # so that empty cells don't cause column shifting
        row_data = {}
        for cell in row.findall('s:c', ns):
            ref = cell.get('r', '')          # e.g. 'E5'
            col_letters = re.match(r'([A-Z]+)', ref).group(1) if ref else None
            if not col_letters:
                continue
            col_idx = col_letter_to_index(col_letters)
            t = cell.get('t', '')
            v_el = cell.find('s:v', ns)
            if v_el is not None:
                if t == 's':
                    row_data[col_idx] = strings[int(v_el.text)]
                else:
                    row_data[col_idx] = v_el.text
        if row_data:
            max_col = max(row_data.keys())
            rows.append([row_data.get(i, '') for i in range(max_col + 1)])
        else:
            rows.append([])
    return rows

def filename_from_url(url):
    if not url:
        return None
    return url.rstrip('/').split('/')[-1]

# ---- Load SharePoint file lists (add the newly uploaded file) ----
sp_file_to_folder = {}

sp_folder_map = {
    'datasheets': 'query(datasheets).iqy.xlsx',
    'instructions': 'query(instructions).iqy.xlsx',
    'images': 'query(prodcutImages.iqy.xlsx',
}

for folder_key, fname in sp_folder_map.items():
    with zipfile.ZipFile(fname) as z:
        ss = get_shared_strings(z)
        sp_rows = get_sheet_data(z, ss)
    name_col = sp_rows[0].index('Name')
    for r in sp_rows[1:]:
        if name_col < len(r) and r[name_col]:
            sp_file_to_folder[r[name_col]] = folder_key

# Files uploaded after the SharePoint exports were taken
sp_file_to_folder['9546_ins_temp_rsz.pdf'] = 'instructions'
sp_file_to_folder['38242_ins_weather_shade.pdf'] = 'instructions'
sp_file_to_folder['Thermowells_NoPrice-v17.pdf'] = 'datasheets'
sp_file_to_folder['Water_Leak_Detector_BB_5Amp_NoPrice.pdf'] = 'datasheets'
sp_file_to_folder['Wireless_BLE_Quantum_TempRH_NoPrice.pdf'] = 'datasheets'
sp_file_to_folder['37-55-CDW.png'] = 'images'
sp_file_to_folder['Clamp_BB2_300pix.png'] = 'images'
sp_file_to_folder['Duct_Avg_BBX.png'] = 'images'
sp_file_to_folder['Immersion_BB2O.png'] = 'images'
sp_file_to_folder['Remote_Sensor_BB4-1.png'] = 'images'
sp_file_to_folder['Remote_Sensor_BB_300pix.png'] = 'images'
sp_file_to_folder['Remote_Sensor_WP_300pix.png'] = 'images'
sp_file_to_folder['Thermobuffer_BBX-_2inch_300pix.png'] = 'images'

print(f"SharePoint file index built: {len(sp_file_to_folder)} files")

# ---- Build URL mapping ----
def map_url(old_url):
    """Convert a bapihvac.com URL to a SharePoint URL, or return as-is if not mappable."""
    if not old_url or 'bapihvac.com' not in old_url:
        return old_url, 'unchanged'

    fname = filename_from_url(old_url)
    if not fname or fname == 'l.pdf':
        return '', 'blanked (invalid URL)'

    # Check which SharePoint folder has this file
    if fname in sp_file_to_folder:
        folder_key = sp_file_to_folder[fname]
        # URL-encode just the filename (spaces, parens, etc.)
        encoded_fname = quote(fname, safe='.-_')
        new_url = f"{SP_FOLDERS[folder_key]}/{encoded_fname}"
        return new_url, f'mapped → {folder_key}'
    else:
        return old_url, 'NOT IN SHAREPOINT - kept as-is'

# ---- Load and rewrite spreadsheet ----
with zipfile.ZipFile('Kele_Contents_BAPI_Final.xlsx') as z:
    strings = get_shared_strings(z)
    rows = get_sheet_data(z, strings)

headers = rows[0]
img_col = headers.index('Product Image URL')
ds_col = headers.index('Product Data Sheet PDF URL')
ins_col = headers.index('Product Installation Instructions URL')

# ---- Write new workbook ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Style for header row
from openpyxl.styles import Font, PatternFill, Alignment
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Track issues
not_mapped = []
blanked = []

# Write header
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Write data rows
for row_idx, row in enumerate(rows[1:], 2):
    # Pad row to full width if needed
    while len(row) < len(headers):
        row.append('')

    new_row = list(row)

    # Map image URL
    if img_col < len(row) and row[img_col]:
        new_url, status = map_url(row[img_col])
        new_row[img_col] = new_url
        if 'NOT IN SHAREPOINT' in status:
            not_mapped.append((row_idx, 'Image', row[img_col]))
        elif 'blanked' in status:
            blanked.append((row_idx, 'Image', row[img_col]))

    # Map datasheet URL
    if ds_col < len(row) and row[ds_col]:
        new_url, status = map_url(row[ds_col])
        new_row[ds_col] = new_url
        if 'NOT IN SHAREPOINT' in status:
            not_mapped.append((row_idx, 'Datasheet', row[ds_col]))
        elif 'blanked' in status:
            blanked.append((row_idx, 'Datasheet', row[ds_col]))

    # Map instruction URL
    if ins_col < len(row) and row[ins_col]:
        new_url, status = map_url(row[ins_col])
        new_row[ins_col] = new_url
        if 'NOT IN SHAREPOINT' in status:
            not_mapped.append((row_idx, 'Instructions', row[ins_col]))

    for col_idx, val in enumerate(new_row, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)

# Auto-size columns (approximate)
col_widths = {1: 20, 2: 30, 3: 55, 4: 55, 5: 90, 6: 90, 7: 90}
for col_num, width in col_widths.items():
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

# Freeze header row
ws.freeze_panes = 'A2'

wb.save('Kele_Contents_BAPI_Updated.xlsx')
print(f"\nSaved: Kele_Contents_BAPI_Updated.xlsx")
print(f"Total data rows: {len(rows)-1}")

# ---- Summary report ----
print(f"\n{'='*55}")
print("RESULTS SUMMARY")
print(f"{'='*55}")

# Count actual mappings
mapped_count = 0
total_url_cells = 0
for row in rows[1:]:
    for col in [img_col, ds_col, ins_col]:
        if col < len(row) and row[col] and 'bapihvac.com' in row[col]:
            total_url_cells += 1
            fname = filename_from_url(row[col])
            if fname and fname in sp_file_to_folder:
                mapped_count += 1

print(f"URL cells processed:        {total_url_cells:,}")
print(f"Successfully mapped to SP:  {mapped_count:,}")
print(f"Kept as-is (not in SP):     {len(not_mapped)}")
print(f"Blanked (invalid URLs):     {len(blanked)}")

if not_mapped:
    # Deduplicate by filename
    seen = set()
    unique_not_mapped = []
    for r, col, url in not_mapped:
        fname = filename_from_url(url)
        if fname not in seen:
            seen.add(fname)
            unique_not_mapped.append(fname)
    print(f"\nUnique files NOT in SharePoint ({len(unique_not_mapped)}):")
    for f in sorted(unique_not_mapped):
        print(f"  ⚠  {f}")

if blanked:
    print(f"\nBlanked invalid URLs:")
    for r, col, url in blanked:
        print(f"  Row {r} [{col}]: {url}")

print(f"\n{'='*55}")
print("✓ Done. Deliver Kele_Contents_BAPI_Updated.xlsx to Kele.")
